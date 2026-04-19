import shutil
import openpyxl
import os

from processor.handlers import *


# ---------------- NORMALIZADORES ----------------

def normalizar_id(valor):
    if not valor:
        return ""
    return "".join(filter(str.isdigit, valor))


def normalizar_nombre(nombre):
    if not nombre:
        return ""
    return nombre.strip().upper()


# ---------------- CLASIFICADOR ----------------

def clasificar_anexo(dte, nit, dui, nombre):

    tipo = dte.get("identificacion", {}).get("tipoDte")

    receptor = dte.get("receptor", {})
    emisor = dte.get("emisor", {})

    receptor_id = normalizar_id(receptor.get("nit"))
    emisor_id = normalizar_id(emisor.get("nit"))

    receptor_nombre = normalizar_nombre(receptor.get("nombre"))
    emisor_nombre = normalizar_nombre(emisor.get("nombre"))

    mi_ids = [normalizar_id(nit), normalizar_id(dui)]
    mi_nombre = normalizar_nombre(nombre)

    soy_receptor = receptor_id in mi_ids or receptor_nombre == mi_nombre
    soy_emisor = emisor_id in mi_ids or emisor_nombre == mi_nombre

    # 🚨 VALIDACIÓN CRÍTICA
    if soy_receptor and soy_emisor:
        return "inconsistente"

    if tipo == "07":
        return 4  # RETENCIONES

    if soy_receptor:
        return 3  # COMPRAS

    if soy_emisor:
        if tipo == "01":
            return 2  # CONSUMIDOR FINAL
        if tipo == "03":
            return 1  # CONTRIBUYENTES

    return None


# ---------------- SERVICIO PRINCIPAL ----------------

def generar_anexos(documentos, nit, dui, nombre, salida_path):

    # 🔥 Ruta robusta de plantilla (clave para Render)
    BASE_DIR = os.path.dirname(os.path.abspath(__file__))
    plantilla = os.path.join(BASE_DIR, "..", "plantilla.xlsm")

    if not os.path.exists(plantilla):
        raise Exception("❌ plantilla.xlsm no encontrada en el servidor")

    shutil.copy(plantilla, salida_path)

    wb = openpyxl.load_workbook(salida_path, keep_vba=True)

    ws_contrib = wb["ANEXO CONTRIBUYENTES"]
    ws_consumidor = wb["ANEXO CONSUMIDOR FINAL"]
    ws_compras = wb["ANEXO DE COMPRAS"]
    ws_retencion = wb["CASILLA 162"]

    # ---------------- PROCESADORES ----------------

    def procesar_retencion(data, fila):

        identificacion = data.get("identificacion", {})
        emisor = data.get("emisor", {})
        cuerpo = data.get("cuerpoDocumento", [{}])[0]

        numero_control = identificacion.get("numeroControl", "")
        partes = numero_control.split("-")

        serie = partes[2] if len(partes) > 2 else ""
        numero = partes[3] if len(partes) > 3 else ""

        valores = [
            emisor.get("nit"),
            formatear_fecha(identificacion.get("fecEmi")),
            "07 Retención",
            serie,
            numero,
            cuerpo.get("montoSujetoGrav", 0),
            cuerpo.get("ivaRetenido", 0),
            emisor.get("nit"),
            7
        ]

        for col, valor in enumerate(valores, start=1):
            ws_retencion.cell(row=fila, column=col).value = valor


    def procesar_contribuyente(data, fila):

        resumen = data.get("resumen", {})
        identificacion = data.get("identificacion", {})
        receptor = data.get("receptor", {})

        iva = ""
        tributos = resumen.get("tributos")

        if tributos:
            for t in tributos:
                if t.get("codigo") == "20":
                    iva = round(t.get("valor", 0), 2)

        valores = [
            formatear_fecha(identificacion.get("fecEmi")),
            "4. DOCUMENTO TRIBUTARIO ELECTRÓNICO (DTE)",
            "03. COMPROBANTE DE CRÉDITO FISCAL",
            identificacion.get("numeroControl"),
            data.get("selloRecibido"),
            identificacion.get("codigoGeneracion"),
            identificacion.get("codigoGeneracion"),
            receptor.get("nit"),
            receptor.get("nombre"),
            vacio_si_cero(resumen.get("totalExenta")),
            vacio_si_cero(resumen.get("totalNoSuj")),
            round(resumen.get("totalGravada", 0), 2),
            iva,
            "",
            "",
            round(resumen.get("montoTotalOperacion", 0), 2),
            "",
            tipo_operacion(identificacion.get("tipoOperacion")),
            "03 Actividades Comerciales",
            1
        ]

        for col, valor in enumerate(valores, start=1):
            ws_contrib.cell(row=fila, column=col).value = valor


    def procesar_consumidor(data, fila):

        resumen = data.get("resumen", {})
        identificacion = data.get("identificacion", {})

        total = round(resumen.get("montoTotalOperacion", 0), 2)

        valores = [
            formatear_fecha(identificacion.get("fecEmi")),
            "4. DOCUMENTO TRIBUTARIO ELECTRÓNICO (DTE)",
            "01. FACTURA",
            identificacion.get("numeroControl"),
            data.get("acuseMH", {}).get("numValidacion"),
            identificacion.get("codigoGeneracion"),
            identificacion.get("codigoGeneracion"),
            "",
            "",
            "",
            0.00,
            0.00,
            0.00,
            total,
            "",
            "",
            "",
            "",
            "",
            total,
            tipo_operacion(identificacion.get("tipoOperacion")),
            "02 Actividades de Servicios",
            2
        ]

        for col, valor in enumerate(valores, start=1):
            ws_consumidor.cell(row=fila, column=col).value = valor


    def procesar_compra(data, fila):

        resumen = data.get("resumen", {})
        identificacion = data.get("identificacion", {})
        emisor = data.get("emisor", {})

        iva = 0
        tributos = resumen.get("tributos")

        if tributos:
            for t in tributos:
                if t.get("codigo") == "20":
                    iva = round(t.get("valor", 0), 2)

        valores = [
            formatear_fecha(identificacion.get("fecEmi")),
            "4. DOCUMENTO TRIBUTARIO ELECTRONICO (DTE)",
            "03. COMPROBANTE DE CRÉDITO FISCAL",
            identificacion.get("numeroControl"),
            formatear_nit(emisor.get("nit")),
            emisor.get("nombre"),
            "",
            "",
            "",
            round(resumen.get("totalGravada", 0), 2),
            "",
            "",
            "",
            iva,
            round(resumen.get("montoTotalOperacion", 0), 2),
            "",
            "1 Gravada",
            "2 Gasto",
            "4 Servicios, Profesiones, Artes y Oficios",
            "2 Gasto de Administración sin Donación",
            3
        ]

        for col, valor in enumerate(valores, start=1):
            ws_compras.cell(row=fila, column=col).value = valor


    # ---------------- ORDENAR DOCUMENTOS ----------------

    docs_ordenados = []

    for dte in documentos:
        fecha = dte.get("identificacion", {}).get("fecEmi")
        numero = dte.get("identificacion", {}).get("numeroControl", "")

        docs_ordenados.append((dte, obtener_fecha_obj(fecha), numero))

    docs_ordenados.sort(key=lambda x: (x[1], x[2]))

    # ---------------- PROCESO ----------------

    fila_contrib = 3
    fila_consumidor = 3
    fila_compras = 3
    fila_retencion = 3

    inconsistencias = []

    for data, _, _ in docs_ordenados:

        anexo = clasificar_anexo(data, nit, dui, nombre)

        if anexo == "inconsistente":
            inconsistencias.append({
                "numero": data.get("identificacion", {}).get("numeroControl"),
                "fecha": data.get("identificacion", {}).get("fecEmi")
            })
            continue

        if anexo == 1:
            procesar_contribuyente(data, fila_contrib)
            fila_contrib += 1

        elif anexo == 2:
            procesar_consumidor(data, fila_consumidor)
            fila_consumidor += 1

        elif anexo == 3:
            procesar_compra(data, fila_compras)
            fila_compras += 1

        elif anexo == 4:
            procesar_retencion(data, fila_retencion)
            fila_retencion += 1

    wb.save(salida_path)

    return salida_path, inconsistencias